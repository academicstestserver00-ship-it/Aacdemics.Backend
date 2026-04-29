"""
Teacher Routes
API endpoints for teacher operations — Firebase Realtime Database
"""

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from typing import List, Optional
from pydantic import BaseModel
from datetime import datetime, timezone, timedelta
from firebase_admin import db
from io import BytesIO
from app.routes.auth import require_teacher, get_current_user, get_user_by_email
from app.database import log_audit_event
from app.utils.rbac import has_permission, normalize_role
import math
import random
import string
import uuid

router = APIRouter(prefix="/api/teacher", tags=["teacher"])

ALL_LANGUAGES = ["python", "c", "cpp", "java"]
IST = timezone(timedelta(hours=5, minutes=30))


# ─── Pydantic Models ──────────────────────────────────────────────────────────

class TestCreate(BaseModel):
    title: str
    description: str
    duration_minutes: int = 60
    is_active: bool = True
    allowed_languages: List[str] = ["python", "c", "cpp", "java"]
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    test_type: str = "invite_only"
    tags: Optional[str] = ""
    anti_paste_enabled: Optional[bool] = True
    tab_switch_enabled: Optional[bool] = True
    tab_switch_limit: Optional[int] = 3
    auto_end_at_end_date: Optional[bool] = True
    negative_marking_enabled: Optional[bool] = False
    negative_marking_marks: Optional[float] = 0.0
    geo_fencing_enabled: Optional[bool] = False
    geo_latitude: Optional[float] = None
    geo_longitude: Optional[float] = None
    geo_radius_meters: Optional[int] = 100


class TestUpdate(BaseModel):
    is_active: Optional[bool] = None
    start_date: Optional[str] = None
    end_date: Optional[str] = None
    test_type: Optional[str] = None
    tags: Optional[str] = None
    title: Optional[str] = None
    description: Optional[str] = None
    duration_minutes: Optional[int] = None
    anti_paste_enabled: Optional[bool] = None
    tab_switch_enabled: Optional[bool] = None
    tab_switch_limit: Optional[int] = None
    auto_end_at_end_date: Optional[bool] = None
    negative_marking_enabled: Optional[bool] = None
    negative_marking_marks: Optional[float] = None
    geo_fencing_enabled: Optional[bool] = None
    geo_latitude: Optional[float] = None
    geo_longitude: Optional[float] = None
    geo_radius_meters: Optional[int] = None


class QuestionCreate(BaseModel):
    test_id: str
    title: str
    description: str
    difficulty: str
    topic: str
    points: int = 10
    time_limit_ms: int = 2000
    question_type: Optional[str] = "coding"  # "coding" or "mcq"
    mcq_options: Optional[List[str]] = None
    mcq_correct_option: Optional[int] = None
    mcq_negative_enabled: Optional[bool] = False
    mcq_negative_marks: Optional[float] = 0.0
    image_url: Optional[str] = None


class QuestionUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    difficulty: Optional[str] = None
    topic: Optional[str] = None
    points: Optional[int] = None
    time_limit_ms: Optional[int] = None
    question_type: Optional[str] = None
    mcq_options: Optional[List[str]] = None
    mcq_correct_option: Optional[int] = None
    mcq_negative_enabled: Optional[bool] = None
    mcq_negative_marks: Optional[float] = None
    image_url: Optional[str] = None


class TestCaseCreate(BaseModel):
    question_id: str
    input: str
    expected_output: str
    is_hidden: bool = False
    points: int = 1


class TestCaseUpdate(BaseModel):
    input: Optional[str] = None
    expected_output: Optional[str] = None
    is_hidden: Optional[bool] = None
    points: Optional[int] = None


class SubmitRequest(BaseModel):
    question_id: str
    test_id: str
    language: Optional[str] = None
    code: Optional[str] = None
    score: Optional[float] = None
    passed: Optional[int] = None
    total: Optional[int] = None
    auto_submit: Optional[bool] = False
    execution_time_ms: Optional[int] = None
    compilation_time_ms: Optional[int] = None
    selected_option: Optional[int] = None


class StartAttemptRequest(BaseModel):
    roll_number: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    accuracy_meters: Optional[float] = None


class ForfeitRequest(BaseModel):
    tab_switches: Optional[int] = None


class TabSwitchLogRequest(BaseModel):
    count: int
    timestamp: Optional[str] = None


class PasteLogRequest(BaseModel):
    count: int
    timestamp: Optional[str] = None


class TeacherAccessGrantRequest(BaseModel):
    email: str
    permission: str = "view"


# ─── Helpers ──────────────────────────────────────────────────────────────────

def parse_languages(lang_str: str) -> List[str]:
    return [l.strip() for l in (lang_str or "python").split(",") if l.strip()]

def format_languages(langs: List[str]) -> str:
    valid = [l.lower() for l in langs if l.lower() in ALL_LANGUAGES]
    return ",".join(valid) if valid else "python"

def normalize_question_type(value: Optional[str]) -> str:
    raw = (value or "coding").strip().lower()
    return "mcq" if raw in {"mcq", "multiple_choice", "multiple-choice", "choice"} else "coding"

def generate_assessment_id() -> str:
    return ''.join(random.choices(string.digits, k=7))

def parse_date(date_str: Optional[str]):
    if not date_str:
        return None
    try:
        dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
        # `datetime-local` values from UI have no timezone; treat them as IST.
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=IST)
        # Store normalized UTC timestamps for consistent comparisons.
        return dt.astimezone(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    except Exception:
        return None

def utcnow_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

def normalize_teacher_permission(value: Optional[str]) -> str:
    normalized = (value or "view").strip().lower()
    if normalized not in {"view", "edit"}:
        raise HTTPException(status_code=400, detail="Permission must be either 'view' or 'edit'.")
    return normalized

def display_language(value: Optional[str]) -> str:
    labels = {
        "python": "Python",
        "py": "Python",
        "c": "C",
        "cpp": "C++",
        "c++": "C++",
        "java": "Java",
        "mcq": "MCQ",
    }
    key = (value or "").strip().lower()
    return labels.get(key, value or "N/A")

def format_duration_label(seconds: Optional[int]) -> str:
    if seconds is None:
        return "N/A"
    total = max(0, int(seconds))
    return f"{total // 60}m {total % 60}s"

def normalize_geo_radius(value: Optional[int]) -> int:
    try:
        radius = int(value or 100)
    except Exception:
        radius = 100
    return max(25, min(radius, 5000))

def has_valid_geofence(test: dict) -> bool:
    if not bool(test.get("geo_fencing_enabled")):
        return False
    try:
        lat = float(test.get("geo_latitude"))
        lng = float(test.get("geo_longitude"))
    except (TypeError, ValueError):
        return False
    return -90 <= lat <= 90 and -180 <= lng <= 180

def calculate_distance_meters(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    earth_radius_m = 6371000
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    d_phi = math.radians(lat2 - lat1)
    d_lambda = math.radians(lng2 - lng1)
    a = (
        math.sin(d_phi / 2) ** 2
        + math.cos(phi1) * math.cos(phi2) * math.sin(d_lambda / 2) ** 2
    )
    return earth_radius_m * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

def ensure_geofence_access(test: dict, data: Optional[StartAttemptRequest]) -> Optional[dict]:
    if not bool(test.get("geo_fencing_enabled")):
        return None
    if not has_valid_geofence(test):
        raise HTTPException(status_code=403, detail="Test location is not configured correctly. Please contact your teacher.")
    if not data or data.latitude is None or data.longitude is None:
        raise HTTPException(status_code=403, detail="Location permission is required to start this geofenced test.")

    try:
        student_lat = float(data.latitude)
        student_lng = float(data.longitude)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Invalid location coordinates.")
    if not (-90 <= student_lat <= 90 and -180 <= student_lng <= 180):
        raise HTTPException(status_code=400, detail="Invalid location coordinates.")

    center_lat = float(test.get("geo_latitude"))
    center_lng = float(test.get("geo_longitude"))
    radius = normalize_geo_radius(test.get("geo_radius_meters"))
    distance = calculate_distance_meters(center_lat, center_lng, student_lat, student_lng)
    accuracy = data.accuracy_meters
    try:
        accuracy_value = float(accuracy) if accuracy is not None else None
    except (TypeError, ValueError):
        accuracy_value = None

    if distance > radius:
        raise HTTPException(
            status_code=403,
            detail=f"You must be within {radius} meters of the assigned test location to start this test.",
        )

    return {
        "latitude": student_lat,
        "longitude": student_lng,
        "accuracy_meters": accuracy_value,
        "distance_meters": round(distance, 2),
        "verified_at": utcnow_iso(),
    }

def get_test_or_404(test_id: str) -> dict:
    test = db.reference(f"/tests/{test_id}").get()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    return test

def _resolve_owner_id(test: dict) -> Optional[str]:
    return test.get("created_by") or test.get("teacher_id")


def get_teacher_permission(test: dict, current_user: dict) -> str:
    user_id = current_user.get("id")
    role = normalize_role(current_user.get("role"))

    if role == "root_superadmin":
        return "owner"

    owner_id = _resolve_owner_id(test)
    if owner_id == user_id:
        return "owner"

    access = test.get("teacher_access") or {}
    entry = access.get(user_id) or {}
    permission = (entry.get("permission") or "").strip().lower()
    if permission == "edit":
        return "edit"
    if permission == "view":
        return "view"
    return "none"

def can_view_test(test: dict, current_user: dict) -> bool:
    return get_teacher_permission(test, current_user) in {"owner", "edit", "view"}

def can_edit_test(test: dict, current_user: dict) -> bool:
    return get_teacher_permission(test, current_user) in {"owner", "edit"}

def can_manage_test_access(test: dict, current_user: dict) -> bool:
    return get_teacher_permission(test, current_user) == "owner"

def ensure_test_view_access(test: dict, current_user: dict):
    if not can_view_test(test, current_user):
        raise HTTPException(status_code=403, detail="You do not have access to this test.")

def ensure_test_edit_access(test: dict, current_user: dict):
    if not can_edit_test(test, current_user):
        raise HTTPException(status_code=403, detail="You do not have edit access to this test.")

def ensure_test_owner_access(test: dict, current_user: dict):
    if not can_manage_test_access(test, current_user):
        raise HTTPException(status_code=403, detail="Only the test owner can manage teacher access.")

def get_question_or_404(question_id: str) -> dict:
    question = db.reference(f"/questions/{question_id}").get()
    if not question:
        raise HTTPException(status_code=404, detail=f"Question {question_id} not found")
    return question

def get_test_case_or_404(test_case_id: str) -> dict:
    test_case = db.reference(f"/test_cases/{test_case_id}").get()
    if not test_case:
        raise HTTPException(status_code=404, detail=f"Test case {test_case_id} not found")
    return test_case

def ensure_session_id(test_id: str, test: dict) -> str:
    session_id = test.get("current_session_id")
    if not session_id:
        session_id = str(uuid.uuid4())
        db.reference(f"/tests/{test_id}").update({"current_session_id": session_id})
        test["current_session_id"] = session_id
    return session_id

# AFTER
def parse_iso_ts(value: Optional[str]):
    if not value:
        return None
    try:
        dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
        # If naive (no timezone), assume IST.
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=IST)
        return dt
    except Exception:
        return None

def format_ist_label(dt: datetime) -> str:
    return dt.astimezone(IST).strftime("%Y-%m-%d %I:%M:%S %p IST")

def is_test_expired(test: dict, now: Optional[datetime] = None) -> bool:
    end_dt = parse_iso_ts(test.get("end_date"))
    if not end_dt:
        return False
    ref = now or datetime.now(timezone.utc)
    return ref >= end_dt

def ensure_test_closed_if_expired(test_id: str, test: dict, now: Optional[datetime] = None) -> dict:
    if not test:
        return test
    auto_end = test.get("auto_end_at_end_date")
    if auto_end is None:
        auto_end = True
    if test.get("is_active") and auto_end and is_test_expired(test, now=now):
        db.reference(f"/tests/{test_id}").update({"is_active": False})
        test["is_active"] = False
    return test

def compute_live_status(test_id: str, test: dict, now: Optional[datetime] = None) -> dict:
    if not test:
        return {
            "test_id": test_id,
            "is_active": False,
            "active_count": 0,
            "total_attempts": 0,
            "forfeited_count": 0,
            "expired_count": 0,
            "submitted_count": 0,
            "active_students": [],
        }
    ref_now = now or datetime.now(timezone.utc)
    session_id = ensure_session_id(test_id, test)
    attempts = db.reference(f"/attempts/{test_id}/{session_id}").get() or {}
    all_users = db.reference("/users").get() or {}
    all_subs = db.reference("/submissions").get() or {}
    submitted_student_ids = {
        s.get("student_id")
        for s in all_subs.values()
        if s.get("test_id") == test_id
        and s.get("session_id") == session_id
        and s.get("student_id")
    }

    active_students = []
    total_attempts = 0
    forfeited_count = 0
    expired_count = 0
    submitted_count = 0

    for student_id, attempt in attempts.items():
        started_at = attempt.get("started_at")
        if not started_at:
            continue
        total_attempts += 1
        duration_seconds = get_effective_duration_seconds(test, started_at=started_at, now=ref_now)
        expired = is_attempt_expired(attempt, duration_seconds)
        forfeited = is_attempt_forfeited(attempt)
        submitted = student_id in submitted_student_ids

        if forfeited:
            forfeited_count += 1
        if expired:
            expired_count += 1
        if submitted:
            submitted_count += 1

        if (not expired) and (not forfeited):
            started_dt = parse_iso_ts(started_at)
            elapsed = (ref_now - started_dt).total_seconds() if started_dt else 0
            remaining = max(0, int(duration_seconds - elapsed))
            user = all_users.get(student_id, {}) if student_id else {}
            active_students.append({
                "student_id": student_id,
                "student_name": user.get("name") or user.get("full_name") or "Unknown",
                "student_email": user.get("email") or "Unknown",
                "started_at": started_at,
                "remaining_seconds": remaining,
            })

    return {
        "test_id": test_id,
        "current_session_id": session_id,
        "is_active": bool(test.get("is_active")),
        "start_date": test.get("start_date"),
        "end_date": test.get("end_date"),
        "auto_end_at_end_date": test.get("auto_end_at_end_date", True),
        "now": ref_now.replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "active_count": len(active_students),
        "total_attempts": total_attempts,
        "forfeited_count": forfeited_count,
        "expired_count": expired_count,
        "submitted_count": submitted_count,
        "active_students": active_students,
        "session_history": test.get("session_history", []),
    }

def get_test_duration_seconds(test: dict) -> int:
    try:
        minutes = int(test.get("duration_minutes", 60))
    except Exception:
        minutes = 60
    return max(1, minutes * 60)

def get_effective_duration_seconds(test: dict, started_at: Optional[str] = None, now: Optional[datetime] = None) -> int:
    """
    Effective duration is the smaller of:
    - configured test duration
    - time remaining until end_date (if set)
    """
    base = get_test_duration_seconds(test)
    end_dt = parse_iso_ts(test.get("end_date"))
    if not end_dt:
        return base
    ref_dt = parse_iso_ts(started_at) if started_at else (now or datetime.now(timezone.utc))
    if not ref_dt:
        return base
    remaining_window = (end_dt - ref_dt).total_seconds()
    return max(1, int(min(base, remaining_window)))

def get_or_create_attempt(test_id: str, session_id: str, student_id: str) -> dict:
    attempt_ref = db.reference(f"/attempts/{test_id}/{session_id}/{student_id}")
    attempt = attempt_ref.get() or {}
    started_at = attempt.get("started_at")
    if not started_at:
        started_at = utcnow_iso()
        attempt = {"started_at": started_at}
        attempt_ref.set(attempt)
    return attempt

def is_attempt_expired(attempt: dict, duration_seconds: int) -> bool:
    started = parse_iso_ts(attempt.get("started_at"))
    if not started:
        return False
    now = datetime.now(timezone.utc)
    return (now - started).total_seconds() > duration_seconds

def is_attempt_forfeited(attempt: dict) -> bool:
    return bool(attempt.get("forfeited"))

def format_test(test_id: str, test: dict, viewer: Optional[dict] = None) -> dict:
    ensure_session_id(test_id, test)
    my_permission = get_teacher_permission(test, viewer) if viewer else None
    return {
        "id": test_id,
        "title": test.get("title"),
        "description": test.get("description"),
        "teacher_id": test.get("teacher_id"),
        "created_by": test.get("created_by") or test.get("teacher_id"),
        "creator_role": normalize_role(test.get("creator_role") or "teacher"),
        "duration_minutes": test.get("duration_minutes", 60),
        "is_active": test.get("is_active", True),
        "allowed_languages": parse_languages(test.get("allowed_languages", "python")),
        "start_date": test.get("start_date"),
        "end_date": test.get("end_date"),
        "test_type": test.get("test_type", "invite_only"),
        "tags": [t.strip() for t in (test.get("tags") or "").split(",") if t.strip()],
        "assessment_id": test.get("assessment_id", ""),
        "created_at": test.get("created_at"),
        "anti_paste_enabled": test.get("anti_paste_enabled", True),
        "tab_switch_enabled": test.get("tab_switch_enabled", True),
        "tab_switch_limit": test.get("tab_switch_limit", 3),
        "auto_end_at_end_date": test.get("auto_end_at_end_date", True),
        "negative_marking_enabled": bool(test.get("negative_marking_enabled")),
        "negative_marking_marks": float(test.get("negative_marking_marks", 0) or 0),
        "geo_fencing_enabled": bool(test.get("geo_fencing_enabled")),
        "geo_latitude": test.get("geo_latitude"),
        "geo_longitude": test.get("geo_longitude"),
        "geo_radius_meters": normalize_geo_radius(test.get("geo_radius_meters")),
        "current_session_id": test.get("current_session_id"),
        "session_history": test.get("session_history", []),
        "my_permission": my_permission,
        "is_owner": my_permission == "owner" if my_permission else None,
    }


def normalize_roll_number(value: Optional[str]) -> Optional[str]:
    text = (value or "").strip()
    return text if text else None


def ensure_can_attempt_test(current_user: dict):
    if not has_permission(current_user.get("role"), "attempt_test"):
        raise HTTPException(status_code=403, detail="Permission 'attempt_test' required.")


# ─── Teacher Routes ───────────────────────────────────────────────────────────

@router.get("/tests")
def get_all_tests(current_user: dict = Depends(require_teacher)):
    all_tests = db.reference("/tests").get() or {}
    now = datetime.now(timezone.utc)
    result = []
    for tid, t in all_tests.items():
        t = ensure_test_closed_if_expired(tid, t, now=now)
        if can_view_test(t, current_user):
            result.append(format_test(tid, t, viewer=current_user))
    return result

@router.get("/tests/{test_id}/live-status")
def get_test_live_status(test_id: str, current_user: dict = Depends(require_teacher)):
    test = get_test_or_404(test_id)
    ensure_test_view_access(test, current_user)
    now = datetime.now(timezone.utc)
    test = ensure_test_closed_if_expired(test_id, test, now=now)
    return compute_live_status(test_id, test, now=now)


@router.post("/tests")
def create_test(test: TestCreate, current_user: dict = Depends(require_teacher)):
    if not has_permission(current_user.get("role"), "create_test"):
        raise HTTPException(status_code=403, detail="Permission 'create_test' required.")

    test_id = str(uuid.uuid4())
    session_id = str(uuid.uuid4())
    tab_switch_limit = int(test.tab_switch_limit or 3)
    if tab_switch_limit < 1:
        tab_switch_limit = 1
    if test.geo_fencing_enabled and (test.geo_latitude is None or test.geo_longitude is None):
        raise HTTPException(status_code=400, detail="Geofencing requires latitude and longitude.")
    test_data = {
        "title": test.title,
        "description": test.description,
        "teacher_id": current_user["id"],
        "created_by": current_user["id"],
        "creator_role": normalize_role(current_user.get("role")),
        "duration_minutes": test.duration_minutes,
        "is_active": test.is_active,
        "allowed_languages": format_languages(test.allowed_languages),
        "start_date": parse_date(test.start_date),
        "end_date": parse_date(test.end_date),
        "test_type": test.test_type or "invite_only",
        "tags": test.tags or "",
        "assessment_id": generate_assessment_id(),
        "created_at": utcnow_iso(),
        "anti_paste_enabled": bool(test.anti_paste_enabled) if test.anti_paste_enabled is not None else True,
        "tab_switch_enabled": bool(test.tab_switch_enabled) if test.tab_switch_enabled is not None else True,
        "tab_switch_limit": tab_switch_limit,
        "auto_end_at_end_date": bool(test.auto_end_at_end_date) if test.auto_end_at_end_date is not None else True,
        "negative_marking_enabled": bool(test.negative_marking_enabled),
        "negative_marking_marks": float(test.negative_marking_marks or 0),
        "geo_fencing_enabled": bool(test.geo_fencing_enabled),
        "geo_latitude": float(test.geo_latitude) if test.geo_latitude is not None else None,
        "geo_longitude": float(test.geo_longitude) if test.geo_longitude is not None else None,
        "geo_radius_meters": normalize_geo_radius(test.geo_radius_meters),
        "current_session_id": session_id,
        "session_history": [],
        "teacher_access": {},
    }
    db.reference(f"/tests/{test_id}").set(test_data)
    log_audit_event(
        user_id=current_user["id"],
        action="create_test",
        resource_id=test_id,
        metadata={"creator_role": normalize_role(current_user.get("role"))},
    )
    return format_test(test_id, test_data, viewer=current_user)


@router.patch("/tests/{test_id}")
def update_test(test_id: str, data: TestUpdate, current_user: dict = Depends(require_teacher)):
    test_ref = db.reference(f"/tests/{test_id}")
    test = test_ref.get()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    ensure_test_edit_access(test, current_user)
    updates = {}
    if data.is_active is not None: updates["is_active"] = data.is_active
    if data.start_date is not None: updates["start_date"] = parse_date(data.start_date)
    if data.end_date is not None: updates["end_date"] = parse_date(data.end_date)
    if data.test_type is not None: updates["test_type"] = data.test_type
    if data.tags is not None: updates["tags"] = data.tags
    if data.title is not None: updates["title"] = data.title
    if data.description is not None: updates["description"] = data.description
    if data.duration_minutes is not None: updates["duration_minutes"] = data.duration_minutes
    if data.anti_paste_enabled is not None: updates["anti_paste_enabled"] = data.anti_paste_enabled
    if data.tab_switch_enabled is not None: updates["tab_switch_enabled"] = data.tab_switch_enabled
    if data.tab_switch_limit is not None:
        limit = int(data.tab_switch_limit)
        updates["tab_switch_limit"] = max(1, limit)
    if data.auto_end_at_end_date is not None:
        updates["auto_end_at_end_date"] = bool(data.auto_end_at_end_date)
    if data.negative_marking_enabled is not None:
        updates["negative_marking_enabled"] = bool(data.negative_marking_enabled)
    if data.negative_marking_marks is not None:
        updates["negative_marking_marks"] = float(data.negative_marking_marks or 0)
    if data.geo_fencing_enabled is not None:
        updates["geo_fencing_enabled"] = bool(data.geo_fencing_enabled)
    if data.geo_latitude is not None:
        updates["geo_latitude"] = float(data.geo_latitude)
    if data.geo_longitude is not None:
        updates["geo_longitude"] = float(data.geo_longitude)
    if data.geo_radius_meters is not None:
        updates["geo_radius_meters"] = normalize_geo_radius(data.geo_radius_meters)
    pending = {**test, **updates}
    if pending.get("geo_fencing_enabled") and not has_valid_geofence(pending):
        raise HTTPException(status_code=400, detail="Geofencing requires valid latitude and longitude.")
    test_ref.update(updates)
    updated = test_ref.get()
    log_audit_event(
        user_id=current_user["id"],
        action="update_test",
        resource_id=test_id,
        metadata={"updated_fields": sorted(list(updates.keys()))},
    )
    return format_test(test_id, updated, viewer=current_user)

@router.post("/tests/{test_id}/start-now")
def start_test_now(test_id: str, current_user: dict = Depends(require_teacher)):
    test_ref = db.reference(f"/tests/{test_id}")
    test = test_ref.get()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    ensure_test_edit_access(test, current_user)
    now = utcnow_iso()
    test_ref.update({"start_date": now, "is_active": True})
    updated = test_ref.get()
    return format_test(test_id, updated, viewer=current_user)

@router.post("/tests/{test_id}/end-now")
def end_test_now(test_id: str, current_user: dict = Depends(require_teacher)):
    test_ref = db.reference(f"/tests/{test_id}")
    test = test_ref.get()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    ensure_test_edit_access(test, current_user)
    now = utcnow_iso()
    test_ref.update({"end_date": now, "is_active": False})
    updated = test_ref.get()
    return format_test(test_id, updated, viewer=current_user)

@router.post("/tests/{test_id}/restart-session")
def restart_test_session(test_id: str, current_user: dict = Depends(require_teacher)):
    test_ref = db.reference(f"/tests/{test_id}")
    test = test_ref.get()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    ensure_test_edit_access(test, current_user)
    now = datetime.now(timezone.utc)
    new_session_id = str(uuid.uuid4())
    prior_session_id = test.get("current_session_id")
    history = list(test.get("session_history") or [])
    if prior_session_id:
        history.append(prior_session_id)
    updates = {
        "current_session_id": new_session_id,
        "start_date": utcnow_iso(),
        "is_active": True,
        "session_history": history,
    }
    end_dt = parse_iso_ts(test.get("end_date"))
    if end_dt and end_dt <= now:
        updates["end_date"] = None
    test_ref.update(updates)
    updated = test_ref.get()
    return format_test(test_id, updated, viewer=current_user)


@router.get("/tests/{test_id}/access")
def list_test_access(test_id: str, current_user: dict = Depends(require_teacher)):
    test = get_test_or_404(test_id)
    ensure_test_owner_access(test, current_user)
    access_map = test.get("teacher_access") or {}
    all_users = db.reference("/users").get() or {}
    collaborators = []
    for teacher_id, access in access_map.items():
        user = all_users.get(teacher_id, {})
        collaborators.append({
            "teacher_id": teacher_id,
            "name": user.get("name") or user.get("full_name") or "Unknown",
            "email": user.get("email") or "",
            "permission": (access.get("permission") or "view").lower(),
            "granted_by": access.get("granted_by"),
            "granted_at": access.get("granted_at"),
        })
    collaborators.sort(key=lambda x: x.get("granted_at") or "", reverse=True)
    return {
        "test_id": test_id,
        "owner_id": test.get("teacher_id"),
        "collaborators": collaborators,
    }


@router.post("/tests/{test_id}/access")
def grant_test_access(
    test_id: str,
    data: TeacherAccessGrantRequest,
    current_user: dict = Depends(require_teacher),
):
    test = get_test_or_404(test_id)
    ensure_test_owner_access(test, current_user)

    email = (data.email or "").strip().lower()
    if not email:
        raise HTTPException(status_code=400, detail="Teacher email is required.")

    teacher = get_user_by_email(email)
    if not teacher or normalize_role(teacher.get("role")) not in {"teacher", "superadmin"}:
        raise HTTPException(status_code=404, detail="Teacher/superadmin not found for this email.")
    if teacher["id"] == test.get("teacher_id"):
        raise HTTPException(status_code=400, detail="Test owner already has full access.")

    permission = normalize_teacher_permission(data.permission)
    access_entry = {
        "permission": permission,
        "granted_by": current_user["id"],
        "granted_at": utcnow_iso(),
    }
    db.reference(f"/tests/{test_id}/teacher_access/{teacher['id']}").set(access_entry)

    return {
        "test_id": test_id,
        "teacher_id": teacher["id"],
        "name": teacher.get("name") or teacher.get("full_name") or "Unknown",
        "email": teacher.get("email"),
        "permission": permission,
        "granted_by": current_user["id"],
        "granted_at": access_entry["granted_at"],
    }


@router.delete("/tests/{test_id}/access/{teacher_id}")
def revoke_test_access(test_id: str, teacher_id: str, current_user: dict = Depends(require_teacher)):
    test = get_test_or_404(test_id)
    ensure_test_owner_access(test, current_user)
    access_map = test.get("teacher_access") or {}
    if teacher_id not in access_map:
        raise HTTPException(status_code=404, detail="Teacher access entry not found.")
    db.reference(f"/tests/{test_id}/teacher_access/{teacher_id}").delete()
    return {"message": "Access revoked successfully."}


@router.delete("/tests/{test_id}")
def delete_test(test_id: str, current_user: dict = Depends(require_teacher)):
    test_ref = db.reference(f"/tests/{test_id}")
    test = test_ref.get()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    ensure_test_owner_access(test, current_user)

    # Delete related questions
    all_questions = db.reference("/questions").get() or {}
    qids = [qid for qid, q in all_questions.items() if q.get("test_id") == test_id]
    for qid in qids:
        db.reference(f"/questions/{qid}").delete()

    # Delete related test cases
    all_tcs = db.reference("/test_cases").get() or {}
    for tcid, tc in all_tcs.items():
        if tc.get("question_id") in qids:
            db.reference(f"/test_cases/{tcid}").delete()

    # Delete related submissions
    all_subs = db.reference("/submissions").get() or {}
    for sid, sub in all_subs.items():
        if sub.get("test_id") == test_id:
            db.reference(f"/submissions/{sid}").delete()

    test_ref.delete()
    log_audit_event(
        user_id=current_user["id"],
        action="delete_test",
        resource_id=test_id,
        metadata={"title": test.get("title")},
    )
    return {"message": "Test deleted"}

@router.get("/test/{test_id}/questions")
def get_test_questions(test_id: str, current_user: dict = Depends(require_teacher)):
    test = get_test_or_404(test_id)
    ensure_test_view_access(test, current_user)
    all_questions = db.reference("/questions").get() or {}
    result = []
    for qid, q in all_questions.items():
        if q.get("test_id") == test_id:
            all_tcs = db.reference("/test_cases").get() or {}
            test_cases = [
                {"id": tcid, "input": tc.get("input"), "expected_output": tc.get("expected_output"),
                 "is_hidden": tc.get("is_hidden", False), "points": tc.get("points", 1)}
                for tcid, tc in all_tcs.items() if tc.get("question_id") == qid
            ]
            q_type = normalize_question_type(q.get("question_type"))
            result.append({
                "id": qid, "title": q.get("title"), "description": q.get("description"),
                "difficulty": q.get("difficulty"), "topic": q.get("topic"),
                "points": q.get("points", 10), "time_limit_ms": q.get("time_limit_ms", 2000),
                "image_url": q.get("image_url"),
                "test_cases_count": len(test_cases), "test_cases": test_cases,
                "question_type": q_type,
                "mcq_options": q.get("mcq_options") if q_type == "mcq" else None,
                "mcq_correct_option": q.get("mcq_correct_option") if q_type == "mcq" else None,
                "mcq_negative_enabled": bool(q.get("mcq_negative_enabled")) if q_type == "mcq" else False,
                "mcq_negative_marks": float(q.get("mcq_negative_marks", 0) or 0) if q_type == "mcq" else 0,
            })
    return result


@router.post("/questions")
def create_question(question: QuestionCreate, current_user: dict = Depends(require_teacher)):
    test = get_test_or_404(question.test_id)
    ensure_test_edit_access(test, current_user)
    qid = str(uuid.uuid4())
    q_type = normalize_question_type(question.question_type)
    q_data = {
        "test_id": question.test_id,
        "title": question.title,
        "description": question.description,
        "difficulty": question.difficulty.upper(),
        "topic": question.topic.upper(),
        "points": question.points,
        "time_limit_ms": question.time_limit_ms,
        "image_url": (question.image_url or "").strip() or None,
        "created_at": utcnow_iso(),
        "question_type": q_type,
    }
    if q_type == "mcq":
        options = [str(o).strip() for o in (question.mcq_options or []) if str(o).strip()]
        if len(options) < 2:
            raise HTTPException(status_code=400, detail="MCQ questions need at least 2 options.")
        if question.mcq_correct_option is None:
            raise HTTPException(status_code=400, detail="MCQ correct option index is required.")
        if question.mcq_correct_option < 0 or question.mcq_correct_option >= len(options):
            raise HTTPException(status_code=400, detail="MCQ correct option index is out of range.")
        q_data["mcq_options"] = options
        q_data["mcq_correct_option"] = int(question.mcq_correct_option)
        q_data["mcq_negative_enabled"] = bool(question.mcq_negative_enabled)
        q_data["mcq_negative_marks"] = float(question.mcq_negative_marks or 0)
    db.reference(f"/questions/{qid}").set(q_data)
    q_data["id"] = qid
    return q_data


@router.put("/questions/{question_id}")
def update_question(question_id: str, question_data: QuestionUpdate, current_user: dict = Depends(require_teacher)):
    q_ref = db.reference(f"/questions/{question_id}")
    q = q_ref.get()
    if not q:
        raise HTTPException(status_code=404, detail=f"Question {question_id} not found")
    test = get_test_or_404(q.get("test_id"))
    ensure_test_edit_access(test, current_user)
    updates = {}
    if question_data.title is not None: updates["title"] = question_data.title
    if question_data.description is not None: updates["description"] = question_data.description
    if question_data.difficulty is not None: updates["difficulty"] = question_data.difficulty.upper()
    if question_data.topic is not None: updates["topic"] = question_data.topic.upper()
    if question_data.points is not None: updates["points"] = question_data.points
    if question_data.time_limit_ms is not None: updates["time_limit_ms"] = question_data.time_limit_ms
    if question_data.image_url is not None:
        updates["image_url"] = (question_data.image_url or "").strip() or None
    if question_data.question_type is not None:
        updates["question_type"] = normalize_question_type(question_data.question_type)
    if question_data.mcq_options is not None or question_data.mcq_correct_option is not None:
        q_type = normalize_question_type(question_data.question_type or q.get("question_type"))
        if q_type != "mcq":
            # If switching away from MCQ, clear MCQ fields
            updates["mcq_options"] = None
            updates["mcq_correct_option"] = None
            updates["mcq_negative_enabled"] = None
            updates["mcq_negative_marks"] = None
        else:
            options = [str(o).strip() for o in (question_data.mcq_options or q.get("mcq_options") or []) if str(o).strip()]
            if len(options) < 2:
                raise HTTPException(status_code=400, detail="MCQ questions need at least 2 options.")
            correct = question_data.mcq_correct_option if question_data.mcq_correct_option is not None else q.get("mcq_correct_option")
            if correct is None or correct < 0 or correct >= len(options):
                raise HTTPException(status_code=400, detail="MCQ correct option index is out of range.")
            updates["mcq_options"] = options
            updates["mcq_correct_option"] = int(correct)
            if question_data.mcq_negative_enabled is not None:
                updates["mcq_negative_enabled"] = bool(question_data.mcq_negative_enabled)
            if question_data.mcq_negative_marks is not None:
                updates["mcq_negative_marks"] = float(question_data.mcq_negative_marks or 0)
    q_ref.update(updates)
    updated = q_ref.get()
    all_tcs = db.reference("/test_cases").get() or {}
    test_cases = [
        {"id": tcid, "input": tc.get("input"), "expected_output": tc.get("expected_output"),
         "is_hidden": tc.get("is_hidden", False), "points": tc.get("points", 1)}
        for tcid, tc in all_tcs.items() if tc.get("question_id") == question_id
    ]
    updated["id"] = question_id
    updated["test_cases"] = test_cases
    return updated


@router.delete("/questions/{question_id}")
def delete_question(question_id: str, current_user: dict = Depends(require_teacher)):
    question = get_question_or_404(question_id)
    test = get_test_or_404(question.get("test_id"))
    ensure_test_edit_access(test, current_user)
    db.reference(f"/questions/{question_id}").delete()
    # Delete related test cases and submissions
    all_tcs = db.reference("/test_cases").get() or {}
    for tcid, tc in all_tcs.items():
        if tc.get("question_id") == question_id:
            db.reference(f"/test_cases/{tcid}").delete()
    all_subs = db.reference("/submissions").get() or {}
    for sid, s in all_subs.items():
        if s.get("question_id") == question_id:
            db.reference(f"/submissions/{sid}").delete()
    return {"message": f"Question {question_id} deleted successfully"}


@router.post("/test-cases")
def create_test_case(test_case: TestCaseCreate, current_user: dict = Depends(require_teacher)):
    question = get_question_or_404(test_case.question_id)
    test = get_test_or_404(question.get("test_id"))
    ensure_test_edit_access(test, current_user)
    tcid = str(uuid.uuid4())
    tc_data = {
        "question_id": test_case.question_id,
        "input": test_case.input,
        "expected_output": test_case.expected_output,
        "is_hidden": test_case.is_hidden,
        "points": test_case.points,
    }
    db.reference(f"/test_cases/{tcid}").set(tc_data)
    tc_data["id"] = tcid
    return tc_data


@router.put("/test-cases/{test_case_id}")
def update_test_case(test_case_id: str, data: TestCaseUpdate, current_user: dict = Depends(require_teacher)):
    tc_ref = db.reference(f"/test_cases/{test_case_id}")
    tc = tc_ref.get()
    if not tc:
        raise HTTPException(status_code=404, detail=f"Test case {test_case_id} not found")
    question = get_question_or_404(tc.get("question_id"))
    test = get_test_or_404(question.get("test_id"))
    ensure_test_edit_access(test, current_user)
    updates = {}
    if data.input is not None: updates["input"] = data.input
    if data.expected_output is not None: updates["expected_output"] = data.expected_output
    if data.is_hidden is not None: updates["is_hidden"] = data.is_hidden
    if data.points is not None: updates["points"] = data.points
    tc_ref.update(updates)
    updated = tc_ref.get()
    updated["id"] = test_case_id
    return updated


@router.delete("/test-cases/{test_case_id}")
def delete_test_case(test_case_id: str, current_user: dict = Depends(require_teacher)):
    test_case = get_test_case_or_404(test_case_id)
    question = get_question_or_404(test_case.get("question_id"))
    test = get_test_or_404(question.get("test_id"))
    ensure_test_edit_access(test, current_user)
    db.reference(f"/test_cases/{test_case_id}").delete()
    return {"message": f"Test case {test_case_id} deleted successfully"}


@router.get("/submissions")
def get_all_submissions(limit: int = 50, current_user: dict = Depends(require_teacher)):
    all_subs = db.reference("/submissions").get() or {}
    all_users = db.reference("/users").get() or {}
    all_questions = db.reference("/questions").get() or {}
    all_tests = db.reference("/tests").get() or {}
    accessible_test_ids = {
        test_id
        for test_id, test_data in all_tests.items()
        if can_view_test(test_data, current_user)
    }
    subs = sorted(all_subs.values(), key=lambda x: x.get("submitted_at", ""), reverse=True)
    enriched = []
    for sub in subs:
        if sub.get("test_id") not in accessible_test_ids:
            continue
        student_id = sub.get("student_id")
        question_id = sub.get("question_id")
        user = all_users.get(student_id, {}) if student_id else {}
        question = all_questions.get(question_id, {}) if question_id else {}
        enriched.append({
            **sub,
            "student_name": user.get("name") or user.get("full_name") or "Unknown",
            "student_email": user.get("email") or "Unknown",
            "student_roll_number": sub.get("roll_number") or user.get("roll_number"),
            "question_title": question.get("title") or "Unknown",
        })
        if len(enriched) >= limit:
            break
    return enriched


@router.get("/analytics/test/{test_id}")
def get_test_analytics(test_id: str, current_user: dict = Depends(require_teacher)):
    test = get_test_or_404(test_id)
    ensure_test_view_access(test, current_user)
    all_questions = db.reference("/questions").get() or {}
    questions = {qid: q for qid, q in all_questions.items() if q.get("test_id") == test_id}
    all_subs = db.reference("/submissions").get() or {}
    session_id = ensure_session_id(test_id, test)
    analytics = {"test_id": test_id, "total_questions": len(questions), "questions": []}
    for qid, q in questions.items():
        subs = [
            s for s in all_subs.values()
            if s.get("question_id") == qid and s.get("session_id") == session_id
        ]
        total = len(subs)
        passed = sum(1 for s in subs if s.get("score") == 100)
        avg = sum(s.get("score", 0) for s in subs) / total if total > 0 else 0
        language_counts = {}
        for s in subs:
            lang = display_language(s.get("language"))
            language_counts[lang] = language_counts.get(lang, 0) + 1
        analytics["questions"].append({
            "question_id": qid, "title": q.get("title"),
            "total_submissions": total, "passed": passed, "failed": total - passed,
            "pass_rate": (passed / total * 100) if total > 0 else 0,
            "average_score": round(avg, 2),
            "languages": language_counts,
            "language": max(language_counts, key=language_counts.get) if language_counts else "N/A",
        })
    return analytics


@router.get("/analytics/test/{test_id}/detailed")
def get_test_analytics_detailed(test_id: str, current_user: dict = Depends(require_teacher)):
    test = get_test_or_404(test_id)
    ensure_test_view_access(test, current_user)
    session_id = ensure_session_id(test_id, test)

    all_questions = db.reference("/questions").get() or {}
    questions = {qid: q for qid, q in all_questions.items() if q.get("test_id") == test_id}
    question_points = {qid: float(q.get("points", 0) or 0) for qid, q in questions.items()}
    question_list = [
        {"id": qid, "title": q.get("title"), "points": q.get("points", 10)}
        for qid, q in questions.items()
    ]

    all_subs = db.reference("/submissions").get() or {}
    all_users = db.reference("/users").get() or {}
    attempts = db.reference(f"/attempts/{test_id}/{session_id}").get() or {}

    duration_seconds = get_test_duration_seconds(test)

    # Aggregate per-question stats
    q_stats = {
        qid: {
            "count": 0,
            "score_sum": 0.0,
            "exec_sum": 0,
            "exec_count": 0,
            "comp_sum": 0,
            "comp_count": 0,
        }
        for qid in questions.keys()
    }

    # Build per-student aggregation
    student_rows = {}
    for sub in all_subs.values():
        if sub.get("test_id") != test_id or sub.get("session_id") != session_id:
            continue
        student_id = sub.get("student_id")
        if not student_id:
            continue
        if student_id not in student_rows:
            user = all_users.get(student_id, {})
            attempt = attempts.get(student_id, {})
            started_at = attempt.get("started_at")
            student_rows[student_id] = {
                "student_id": student_id,
                "student_name": user.get("name") or user.get("full_name") or "Unknown",
                "student_email": user.get("email") or "Unknown",
                "student_roll_number": attempt.get("roll_number") or user.get("roll_number"),
                "started_at": started_at,
                "first_submitted_at": None,
                "deadline_at": None,
                "overall_score": 0,
                "overall_points": 0,
                "overall_submission_time_seconds": None,
                "overall_submitted_at": None,
                "test_submitted": bool(attempt.get("submitted")),
                "test_submitted_at": attempt.get("submitted_at"),
                "tab_switches": attempt.get("tab_switches", 0),
                "paste_count": attempt.get("paste_count", 0),
                "questions": {},
            }
            if started_at:
                started_dt = parse_iso_ts(started_at)
                if started_dt:
                    effective_duration = get_effective_duration_seconds(test, started_at=started_at)
                    student_rows[student_id]["deadline_at"] = (
                        started_dt + timedelta(seconds=effective_duration)
                    ).replace(microsecond=0).isoformat().replace("+00:00", "Z")

        row = student_rows[student_id]
        if not row.get("student_roll_number"):
            row["student_roll_number"] = sub.get("roll_number")
        qid = sub.get("question_id")
        if qid:
            row["questions"][qid] = {
                "question_id": qid,
                "submitted_at": sub.get("submitted_at"),
                "score": sub.get("score", 0),
                "passed": sub.get("passed", 0),
                "total": sub.get("total", 0),
                "points": question_points.get(qid, 0),
                "points_earned": None,
                "language": sub.get("language"),
                "auto_submit": bool(sub.get("auto_submit", False)),
                "execution_time_ms": sub.get("execution_time_ms"),
                "compilation_time_ms": sub.get("compilation_time_ms"),
            }

        # Per-question aggregates
        if qid in q_stats:
            qs = q_stats[qid]
            qs["count"] += 1
            qs["score_sum"] += float(sub.get("score", 0) or 0)
            exec_ms = sub.get("execution_time_ms")
            if exec_ms is not None:
                qs["exec_sum"] += int(exec_ms)
                qs["exec_count"] += 1
            comp_ms = sub.get("compilation_time_ms")
            if comp_ms is not None:
                qs["comp_sum"] += int(comp_ms)
                qs["comp_count"] += 1

        submitted_at = sub.get("submitted_at")
        if submitted_at:
            if (row["first_submitted_at"] is None) or (submitted_at < row["first_submitted_at"]):
                row["first_submitted_at"] = submitted_at
            if (row["overall_submitted_at"] is None) or (submitted_at > row["overall_submitted_at"]):
                row["overall_submitted_at"] = submitted_at

    # Fill missing students with attempts but no submissions
    for student_id, attempt in attempts.items():
        if student_id in student_rows:
            continue
        user = all_users.get(student_id, {})
        started_at = attempt.get("started_at")
        student_rows[student_id] = {
            "student_id": student_id,
            "student_name": user.get("name") or user.get("full_name") or "Unknown",
            "student_email": user.get("email") or "Unknown",
            "student_roll_number": attempt.get("roll_number") or user.get("roll_number"),
            "started_at": started_at,
            "first_submitted_at": None,
            "deadline_at": None,
            "overall_score": 0,
            "overall_points": 0,
            "overall_submission_time_seconds": None,
            "overall_submitted_at": None,
            "test_submitted": bool(attempt.get("submitted")),
            "test_submitted_at": attempt.get("submitted_at"),
            "tab_switches": attempt.get("tab_switches", 0),
            "paste_count": attempt.get("paste_count", 0),
            "questions": {},
        }
        if started_at:
            started_dt = parse_iso_ts(started_at)
            if started_dt:
                effective_duration = get_effective_duration_seconds(test, started_at=started_at)
                student_rows[student_id]["deadline_at"] = (
                    started_dt + timedelta(seconds=effective_duration)
                ).replace(microsecond=0).isoformat().replace("+00:00", "Z")

    # Compute overall score and submission time per student
    total_points = sum(int(q.get("points", 0) or 0) for q in question_list)
    for row in student_rows.values():
        earned_points = 0.0
        if total_points > 0:
            for q in question_list:
                qid = q["id"]
                q_points = float(q.get("points", 0) or 0)
                q_sub = row["questions"].get(qid)
                q_score = float(q_sub.get("score", 0)) if q_sub else 0.0
                points_earned = (q_score / 100.0) * q_points
                earned_points += points_earned
                if q_sub is not None:
                    q_sub["points_earned"] = round(points_earned, 2)
        row["overall_points"] = round(earned_points, 2)
        if total_points > 0:
            row["overall_score"] = round((earned_points / total_points) * 100.0, 2)
        if row["overall_submitted_at"]:
            started_source = row["started_at"] or row.get("first_submitted_at")
            started_dt = parse_iso_ts(started_source)
            submitted_dt = parse_iso_ts(row["overall_submitted_at"])
            if started_dt and submitted_dt:
                row["overall_submission_time_seconds"] = max(0, int((submitted_dt - started_dt).total_seconds()))

        # Aggregate execution/compilation time across submissions
        total_exec = 0
        total_comp = 0
        has_exec = False
        has_comp = False
        for qid, qsub in row["questions"].items():
            exec_ms = qsub.get("execution_time_ms")
            comp_ms = qsub.get("compilation_time_ms")
            if exec_ms is not None:
                total_exec += int(exec_ms)
                has_exec = True
            if comp_ms is not None:
                total_comp += int(comp_ms)
                has_comp = True
        row["total_execution_time_ms"] = total_exec if has_exec else None
        row["total_compilation_time_ms"] = total_comp if has_comp else None
        language_counts = {}
        for qsub in row["questions"].values():
            lang = display_language(qsub.get("language"))
            if lang == "MCQ":
                continue
            language_counts[lang] = language_counts.get(lang, 0) + 1
        row["primary_language"] = max(language_counts, key=language_counts.get) if language_counts else "N/A"

    # Sort by overall_submitted_at desc, then name
    sorted_rows = sorted(
        student_rows.values(),
        key=lambda r: (r.get("overall_submitted_at") or "", r.get("student_name") or ""),
        reverse=True,
    )

    # Attach per-question averages
    enriched_questions = []
    for q in question_list:
        qid = q["id"]
        stats = q_stats.get(qid)
        if stats and stats["count"] > 0:
            avg_score = round(stats["score_sum"] / stats["count"], 2)
        else:
            avg_score = 0
        avg_exec = (
            int(stats["exec_sum"] / stats["exec_count"])
            if stats and stats["exec_count"] > 0
            else None
        )
        avg_comp = (
            int(stats["comp_sum"] / stats["comp_count"])
            if stats and stats["comp_count"] > 0
            else None
        )
        enriched_questions.append({
            **q,
            "average_score": avg_score,
            "avg_execution_time_ms": avg_exec,
            "avg_compilation_time_ms": avg_comp,
        })

    return {
        "test_id": test_id,
        "test_title": test.get("title"),
        "duration_minutes": test.get("duration_minutes", 60),
        "questions": enriched_questions,
        "students": sorted_rows,
    }


# ─── Student Routes ───────────────────────────────────────────────────────────

def _build_teacher_report_workbook(report: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Report"

    dark_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    thin = Side(style="thin", color="B7C9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    bold = Font(bold=True)

    for col, width in {"A": 12, "B": 30, "C": 32, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18}.items():
        ws.column_dimensions[col].width = width

    students = report.get("students") or []
    questions = report.get("questions") or []
    total_marks = sum(float(q.get("points", 0) or 0) for q in questions)
    submitted_students = [s for s in students if s.get("overall_submitted_at") or s.get("test_submitted")]
    average_score = round(sum(float(s.get("overall_score", 0) or 0) for s in students) / len(students), 2) if students else 0

    ws.merge_cells("A1:H1")
    ws["A1"] = "SlashCoder"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].fill = dark_fill
    ws["A1"].alignment = center
    ws.merge_cells("A2:H2")
    ws["A2"] = "Test Performance Report"
    ws["A2"].font = Font(bold=True, size=14)
    ws["A2"].alignment = center
    ws["A4"], ws["B4"] = "Test Name", report.get("test_title") or "Unknown Test"
    ws["D4"], ws["E4"] = "Date", utcnow_iso()[:10]

    ws.merge_cells("A6:H6")
    ws["A6"] = "Overall Summary"
    ws["A6"].font = bold
    ws["A6"].fill = section_fill
    for idx, (label, value) in enumerate([
        ("Total Marks", total_marks),
        ("Average Percentage", f"{average_score}%"),
        ("Total Students", len(students)),
        ("Submitted Students", len(submitted_students)),
        ("Total Questions", len(questions)),
        ("Duration", f"{report.get('duration_minutes', 0)} minutes"),
    ], start=7):
        ws.cell(idx, 1, label)
        ws.cell(idx, 2, value)

    student_start = 15
    ws.merge_cells(start_row=student_start, start_column=1, end_row=student_start, end_column=8)
    ws.cell(student_start, 1, "Student Details")
    ws.cell(student_start, 1).font = bold
    ws.cell(student_start, 1).fill = section_fill
    for col, header in enumerate(["Rank", "Name", "Email / ID", "Language Used", "Total Marks", "Percentage", "Time Taken", "Attempted"], start=1):
        cell = ws.cell(student_start + 1, col, header)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center

    ranked_students = sorted(students, key=lambda s: float(s.get("overall_score", 0) or 0), reverse=True)
    for row_idx, student in enumerate(ranked_students, start=student_start + 2):
        q_entries = student.get("questions") or {}
        values = [
            row_idx - student_start - 1,
            student.get("student_name") or "Unknown",
            student.get("student_email") or "Unknown",
            student.get("primary_language") or "N/A",
            student.get("overall_points", 0),
            f"{student.get('overall_score', 0)}%",
            format_duration_label(student.get("overall_submission_time_seconds")),
            len(q_entries),
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col, value)

    question_start = student_start + 4 + len(ranked_students)
    ws.merge_cells(start_row=question_start, start_column=1, end_row=question_start, end_column=8)
    ws.cell(question_start, 1, "Question-wise Table")
    ws.cell(question_start, 1).font = bold
    ws.cell(question_start, 1).fill = section_fill
    for col, header in enumerate(["Q No", "Question Name", "Submissions", "Passed", "Failed", "Avg Marks %", "Language", "Avg Time"], start=1):
        cell = ws.cell(question_start + 1, col, header)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = center

    total_test_cases = 0
    passed_test_cases = 0
    for row_idx, question in enumerate(questions, start=question_start + 2):
        qid = question.get("id")
        q_subs = [s.get("questions", {}).get(qid) for s in students if s.get("questions", {}).get(qid)]
        passed = sum(1 for q in q_subs if float(q.get("score", 0) or 0) >= 100)
        total = len(q_subs)
        total_test_cases += sum(int(float(q.get("total", 0) or 0)) for q in q_subs)
        passed_test_cases += sum(int(float(q.get("passed", 0) or 0)) for q in q_subs)
        values = [
            row_idx - question_start - 1,
            question.get("title") or "Question",
            total,
            passed,
            max(0, total - passed),
            question.get("average_score", 0),
            question.get("language") or "N/A",
            f"{question.get('avg_execution_time_ms')}ms" if question.get("avg_execution_time_ms") is not None else "N/A",
        ]
        for col, value in enumerate(values, start=1):
            ws.cell(row_idx, col, value)

    tc_start = question_start + 4 + len(questions)
    ws.merge_cells(start_row=tc_start, start_column=1, end_row=tc_start, end_column=2)
    ws.cell(tc_start, 1, "Test Case Summary")
    ws.cell(tc_start, 1).font = bold
    ws.cell(tc_start, 1).fill = section_fill
    ws.cell(tc_start + 1, 1, "Metric")
    ws.cell(tc_start + 1, 2, "Value")
    ws.cell(tc_start + 1, 1).font = bold
    ws.cell(tc_start + 1, 2).font = bold
    ws.cell(tc_start + 1, 1).fill = header_fill
    ws.cell(tc_start + 1, 2).fill = header_fill
    for idx, (metric, value) in enumerate([
        ("Total Test Cases", total_test_cases),
        ("Passed", passed_test_cases),
        ("Failed", max(0, total_test_cases - passed_test_cases)),
    ], start=tc_start + 2):
        ws.cell(idx, 1, metric)
        ws.cell(idx, 2, value)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = border
            cell.alignment = left if cell.column in {2, 3} else center
    ws.freeze_panes = "A17"
    return wb


@router.get("/test-report/{test_id}")
def download_teacher_test_report(test_id: str, current_user: dict = Depends(require_teacher)):
    report = get_test_analytics_detailed(test_id, current_user)
    workbook = _build_teacher_report_workbook(report)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="SlashCoder_Test_Report.xlsx"'},
    )


@router.get("/student/tests", tags=["student"])
def get_available_tests(current_user: dict = Depends(get_current_user)):
    ensure_can_attempt_test(current_user)
    all_tests = db.reference("/tests").get() or {}
    all_questions = db.reference("/questions").get() or {}
    result = []
    now = datetime.now(timezone.utc)
    for tid, t in all_tests.items():
        t = ensure_test_closed_if_expired(tid, t, now=now)
        start_dt = parse_iso_ts(t.get("start_date"))
        end_dt = parse_iso_ts(t.get("end_date"))
        if t.get("is_active") and (not start_dt or now >= start_dt) and (not end_dt or now < end_dt):
            q_count = sum(1 for q in all_questions.values() if q.get("test_id") == tid)
            result.append({
                "id": tid,
                "title": t.get("title"),
                "description": t.get("description"),
                "duration_minutes": t.get("duration_minutes", 60),
                "question_count": q_count,
                "allowed_languages": parse_languages(t.get("allowed_languages", "python")),
                "assessment_id": t.get("assessment_id", ""),  # ← ADD THIS
                "created_at": t.get("created_at"),
                "anti_paste_enabled": t.get("anti_paste_enabled", True),
                "tab_switch_enabled": t.get("tab_switch_enabled", True),
                "tab_switch_limit": t.get("tab_switch_limit", 3),
                "negative_marking_enabled": bool(t.get("negative_marking_enabled")),
                "negative_marking_marks": float(t.get("negative_marking_marks", 0) or 0),
                "geo_fencing_enabled": bool(t.get("geo_fencing_enabled")),
                "geo_radius_meters": normalize_geo_radius(t.get("geo_radius_meters")),
            })
    return result


@router.get("/student/test/{test_id}/questions", tags=["student"])
def get_test_questions_for_student(test_id: str, current_user: dict = Depends(get_current_user)):
    ensure_can_attempt_test(current_user)
    test = db.reference(f"/tests/{test_id}").get()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")

    all_questions = db.reference("/questions").get() or {}
    all_tcs = db.reference("/test_cases").get() or {}

    result = []
    for qid, q in all_questions.items():
        if q.get("test_id") == test_id:
            q_type = normalize_question_type(q.get("question_type"))
            test_cases = [
                {
                    "id": tcid,
                    "input": tc.get("input"),
                    "expected_output": tc.get("expected_output"),
                    "is_hidden": tc.get("is_hidden", False),
                    "points": tc.get("points", 1),
                }
                for tcid, tc in all_tcs.items()
                if tc.get("question_id") == qid
            ]
            result.append({
                "id": qid,
                "title": q.get("title"),
                "description": q.get("description"),
                "difficulty": q.get("difficulty"),
                "topic": q.get("topic"),
                "points": q.get("points", 10),
                "time_limit_ms": q.get("time_limit_ms", 2000),
                "image_url": q.get("image_url"),
                "allowed_languages": parse_languages(test.get("allowed_languages", "python")),
                "test_cases": test_cases if q_type == "coding" else [],
                "question_type": q_type,
                "mcq_options": q.get("mcq_options") if q_type == "mcq" else None,
                "mcq_negative_enabled": bool(q.get("mcq_negative_enabled")) if q_type == "mcq" else False,
                "mcq_negative_marks": float(q.get("mcq_negative_marks", 0) or 0) if q_type == "mcq" else 0,
            })

    return result


@router.get("/student/test/lookup/{code}", tags=["student"])
def lookup_test(code: str, current_user: dict = Depends(get_current_user)):
    ensure_can_attempt_test(current_user)
    all_tests = db.reference("/tests").get() or {}
    all_questions = db.reference("/questions").get() or {}
    now = datetime.now(timezone.utc)
    for tid, t in all_tests.items():
        if tid == code or t.get("assessment_id") == code:
            t = ensure_test_closed_if_expired(tid, t, now=now)
            q_count = sum(1 for q in all_questions.values() if q.get("test_id") == tid)
            return {
                "id": tid,
                "title": t.get("title"),
                "description": t.get("description"),
                "duration_minutes": t.get("duration_minutes", 60),
                "question_count": q_count,
                "allowed_languages": parse_languages(t.get("allowed_languages", "python")),
                "assessment_id": t.get("assessment_id", ""),
                "created_at": t.get("created_at"),
                "anti_paste_enabled": t.get("anti_paste_enabled", True),
                "tab_switch_enabled": t.get("tab_switch_enabled", True),
                "tab_switch_limit": t.get("tab_switch_limit", 3),
                "negative_marking_enabled": bool(t.get("negative_marking_enabled")),
                "negative_marking_marks": float(t.get("negative_marking_marks", 0) or 0),
                "geo_fencing_enabled": bool(t.get("geo_fencing_enabled")),
                "geo_radius_meters": normalize_geo_radius(t.get("geo_radius_meters")),
            }
    raise HTTPException(status_code=404, detail="Test not found")


@router.get("/student/test/{test_id}/submissions", tags=["student"])
def get_student_test_submissions(test_id: str, current_user: dict = Depends(get_current_user)):
    ensure_can_attempt_test(current_user)
    student_id = current_user["id"]
    test = db.reference(f"/tests/{test_id}").get()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    session_id = ensure_session_id(test_id, test)
    all_subs = db.reference("/submissions").get() or {}
    result = []
    for sub_id, sub in all_subs.items():
        if sub.get("student_id") != student_id:
            continue
        if sub.get("test_id") != test_id:
            continue
        if sub.get("session_id") != session_id:
            continue
        result.append({
            "submission_id": sub_id,
            "question_id": sub.get("question_id"),
            "test_id": sub.get("test_id"),
            "session_id": sub.get("session_id"),
            "roll_number": sub.get("roll_number"),
            "language": sub.get("language"),
            "code": sub.get("code"),
            "score": sub.get("score", 0),
            "passed": sub.get("passed", 0),
            "total": sub.get("total", 0),
            "submitted_at": sub.get("submitted_at"),
            "selected_option": sub.get("selected_option"),
            "question_type": sub.get("question_type"),
        })
    result.sort(key=lambda x: x.get("submitted_at") or "", reverse=True)
    return result


@router.post("/student/test/{test_id}/start", tags=["student"])
def start_test_attempt(
    test_id: str,
    data: Optional[StartAttemptRequest] = None,
    current_user: dict = Depends(get_current_user),
):
    ensure_can_attempt_test(current_user)
    test = db.reference(f"/tests/{test_id}").get()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    now = datetime.now(timezone.utc)
    test = ensure_test_closed_if_expired(test_id, test, now=now)
    session_id = ensure_session_id(test_id, test)
    start_dt = parse_iso_ts(test.get("start_date"))
    end_dt = parse_iso_ts(test.get("end_date"))

    if start_dt and now < start_dt:
        raise HTTPException(
            status_code=403,
            detail=f"Test entry opens at {format_ist_label(start_dt)}"
        )
    geo_verification = ensure_geofence_access(test, data)

    attempt_ref = db.reference(f"/attempts/{test_id}/{session_id}/{current_user['id']}")
    attempt = attempt_ref.get() or {}
    started_at = attempt.get("started_at")
    existing_roll = normalize_roll_number(current_user.get("roll_number"))
    provided_roll = normalize_roll_number(data.roll_number if data else None)
    effective_roll = provided_roll or existing_roll

    if not effective_roll:
        raise HTTPException(status_code=400, detail="University roll number is required before starting the test.")

    if provided_roll and provided_roll != existing_roll:
        db.reference(f"/users/{current_user['id']}").update({"roll_number": provided_roll})
        current_user["roll_number"] = provided_roll

    if end_dt and now > end_dt and not started_at:
        raise HTTPException(status_code=403, detail="Test entry window has closed.")

    if not started_at:
        # New attempt starts now, but duration may be reduced by end_date
        duration_seconds = get_effective_duration_seconds(test, started_at=None, now=now)
        if end_dt and (end_dt - now).total_seconds() <= 0:
            raise HTTPException(status_code=403, detail="Test entry window has closed.")
        attempt = {"started_at": utcnow_iso(), "roll_number": effective_roll}
        if geo_verification:
            attempt["geo_verification"] = geo_verification
        attempt_ref.set(attempt)
        started_at = attempt.get("started_at")
        elapsed = 0
    else:
        update_existing = {}
        if attempt.get("roll_number") != effective_roll:
            update_existing["roll_number"] = effective_roll
            attempt["roll_number"] = effective_roll
        if geo_verification:
            update_existing["geo_verification"] = geo_verification
            attempt["geo_verification"] = geo_verification
        if update_existing:
            attempt_ref.update(update_existing)
        duration_seconds = get_effective_duration_seconds(test, started_at=started_at, now=now)
        started = parse_iso_ts(started_at)
        elapsed = (now - started).total_seconds() if started else 0

    remaining = max(0, int(duration_seconds - elapsed))
    return {
        "test_id": test_id,
        "started_at": started_at,
        "duration_seconds": duration_seconds,
        "remaining_seconds": remaining,
        "expired": remaining <= 0,
        "forfeited": is_attempt_forfeited(attempt),
        "roll_number": effective_roll,
        "submitted": bool(attempt.get("submitted")),
        "submitted_at": attempt.get("submitted_at"),
    }


@router.post("/student/test/{test_id}/tab-switch", tags=["student"])
def log_tab_switch(
    test_id: str,
    data: TabSwitchLogRequest,
    current_user: dict = Depends(get_current_user),
):
    ensure_can_attempt_test(current_user)
    test = db.reference(f"/tests/{test_id}").get()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    session_id = ensure_session_id(test_id, test)

    attempt_ref = db.reference(f"/attempts/{test_id}/{session_id}/{current_user['id']}")
    attempt = attempt_ref.get() or {}
    if attempt.get("forfeited"):
        raise HTTPException(status_code=403, detail="Test forfeited.")

    ts = data.timestamp or utcnow_iso()
    event = {"count": data.count, "timestamp": ts}
    db.reference(f"/attempts/{test_id}/{session_id}/{current_user['id']}/tab_switch_events").push(event)
    attempt_ref.update({"tab_switches": data.count, "last_tab_switch_at": ts})
    return {"logged": True, "count": data.count, "timestamp": ts}


@router.post("/student/test/{test_id}/paste", tags=["student"])
def log_paste(
    test_id: str,
    data: PasteLogRequest,
    current_user: dict = Depends(get_current_user),
):
    ensure_can_attempt_test(current_user)
    test = db.reference(f"/tests/{test_id}").get()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    session_id = ensure_session_id(test_id, test)

    attempt_ref = db.reference(f"/attempts/{test_id}/{session_id}/{current_user['id']}")
    attempt = attempt_ref.get() or {}
    if attempt.get("forfeited"):
        raise HTTPException(status_code=403, detail="Test forfeited.")

    ts = data.timestamp or utcnow_iso()
    event = {"count": data.count, "timestamp": ts}
    db.reference(f"/attempts/{test_id}/{session_id}/{current_user['id']}/paste_events").push(event)
    attempt_ref.update({"paste_count": data.count, "last_paste_at": ts})
    return {"logged": True, "count": data.count, "timestamp": ts}


@router.post("/student/test/{test_id}/forfeit", tags=["student"])
def forfeit_test_attempt(
    test_id: str,
    data: ForfeitRequest,
    current_user: dict = Depends(get_current_user),
):
    ensure_can_attempt_test(current_user)
    test = db.reference(f"/tests/{test_id}").get()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    session_id = ensure_session_id(test_id, test)

    attempt_ref = db.reference(f"/attempts/{test_id}/{session_id}/{current_user['id']}")
    attempt = attempt_ref.get() or {}
    if attempt.get("forfeited"):
        return {
            "message": "Test already forfeited.",
            "forfeited": True,
            "forfeited_at": attempt.get("forfeited_at"),
            "tab_switches": attempt.get("tab_switches"),
        }

    updates = {
        "forfeited": True,
        "forfeited_at": utcnow_iso(),
    }
    if data.tab_switches is not None:
        updates["tab_switches"] = data.tab_switches

    attempt_ref.update(updates)
    return {
        "message": "Test forfeited.",
        "forfeited": True,
        "forfeited_at": updates["forfeited_at"],
        "tab_switches": updates.get("tab_switches"),
    }


@router.post("/student/test/{test_id}/submit-test", tags=["student"])
def submit_test(test_id: str, current_user: dict = Depends(get_current_user)):
    ensure_can_attempt_test(current_user)
    test = db.reference(f"/tests/{test_id}").get()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    session_id = ensure_session_id(test_id, test)
    attempt_ref = db.reference(f"/attempts/{test_id}/{session_id}/{current_user['id']}")
    attempt = attempt_ref.get() or {}
    if not attempt.get("started_at"):
        raise HTTPException(status_code=400, detail="Test attempt has not started.")
    if is_attempt_forfeited(attempt):
        raise HTTPException(status_code=403, detail="Test forfeited due to excessive tab switching.")
    if attempt.get("submitted"):
        return {
            "message": "Your test response has been recorded. See you soon.",
            "submitted": True,
            "submitted_at": attempt.get("submitted_at"),
        }

    submitted_at = utcnow_iso()
    attempt_ref.update({"submitted": True, "submitted_at": submitted_at})
    return {
        "message": "Your test response has been recorded. See you soon.",
        "submitted": True,
        "submitted_at": submitted_at,
    }


@router.post("/student/submit", tags=["student"])
def submit_solution(data: SubmitRequest, current_user: dict = Depends(get_current_user)):
    ensure_can_attempt_test(current_user)
    question = db.reference(f"/questions/{data.question_id}").get()
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")
    if question.get("test_id") != data.test_id:
        raise HTTPException(status_code=400, detail="Question does not belong to this test")

    test = db.reference(f"/tests/{data.test_id}").get()
    if not test:
        raise HTTPException(status_code=404, detail="Test not found")
    session_id = ensure_session_id(data.test_id, test)

    # Enforce one submission per question per student
    all_subs = db.reference("/submissions").get() or {}
    for sub in all_subs.values():
        if (
            sub.get("student_id") == current_user["id"]
            and sub.get("question_id") == data.question_id
            and sub.get("test_id") == data.test_id
            and sub.get("session_id") == session_id
        ):
            raise HTTPException(status_code=409, detail="Already submitted for this question")

    # Enforce global test timer (allow small grace for auto-submit)
    attempt_ref = db.reference(f"/attempts/{data.test_id}/{session_id}/{current_user['id']}")
    attempt = attempt_ref.get() or {}
    if not attempt.get("started_at"):
        raise HTTPException(status_code=400, detail="Test attempt has not started.")
    roll_number = normalize_roll_number(attempt.get("roll_number") or current_user.get("roll_number"))
    if not roll_number:
        raise HTTPException(status_code=400, detail="University roll number is required before submitting.")
    if not attempt.get("roll_number"):
        attempt_ref.update({"roll_number": roll_number})
    if attempt.get("submitted"):
        raise HTTPException(status_code=409, detail="Test already submitted.")
    if is_attempt_forfeited(attempt):
        raise HTTPException(status_code=403, detail="Test forfeited due to excessive tab switching.")
    duration_seconds = get_effective_duration_seconds(test, started_at=attempt.get("started_at"))
    if is_attempt_expired(attempt, duration_seconds):
        if not data.auto_submit:
            raise HTTPException(status_code=403, detail="Test time is over")
        # Allow auto-submit within a small grace window after expiry
        started = parse_iso_ts(attempt.get("started_at"))
        if not started:
            raise HTTPException(status_code=403, detail="Test time is over")
        now = datetime.now(timezone.utc)
        end_time = started + timedelta(seconds=duration_seconds)
        if (now - end_time).total_seconds() > 5:
            raise HTTPException(status_code=403, detail="Test time is over")

    q_type = normalize_question_type(question.get("question_type"))
    if q_type == "mcq":
        options = question.get("mcq_options") or []
        correct = question.get("mcq_correct_option")
        selected = data.selected_option
        negative_enabled = bool(question.get("mcq_negative_enabled"))
        negative_marks = float(question.get("mcq_negative_marks", 0) or 0)
        is_correct = selected is not None and correct is not None and int(selected) == int(correct)
        if selected is None:
            score = 0.0
            passed = 0
            total = 1
        elif is_correct:
            score = 100.0
            passed = 1
            total = 1
        else:
            if negative_enabled and negative_marks > 0:
                points = float(question.get("points", 10) or 10)
                score = -round((negative_marks / points) * 100.0, 2) if points > 0 else 0.0
            else:
                score = 0.0
            passed = 0
            total = 1
        language = "mcq"
        code = None
    else:
        if data.score is None or data.passed is None or data.total is None:
            raise HTTPException(status_code=400, detail="Missing score for coding submission.")
        score = data.score
        passed = data.passed
        total = data.total
        language = data.language
        code = data.code

        # Apply test-level negative marking for coding questions (HackerEarth style)
        if bool(test.get("negative_marking_enabled")) and (score is not None) and score < 100:
            neg_marks = float(test.get("negative_marking_marks", 0) or 0)
            if neg_marks > 0:
                q_points = float(question.get("points", 10) or 10)
                if q_points > 0:
                    score = round(score - (neg_marks / q_points) * 100.0, 2)

    sub_id = str(uuid.uuid4())
    submission = {
        "student_id": current_user["id"],
        "roll_number": roll_number,
        "question_id": data.question_id,
        "test_id": data.test_id,
        "session_id": session_id,
        "language": language,
        "code": code,
        "score": score,
        "passed": passed,
        "total": total,
        "auto_submit": bool(data.auto_submit),
        "execution_time_ms": data.execution_time_ms if q_type != "mcq" else None,
        "compilation_time_ms": data.compilation_time_ms if q_type != "mcq" else None,
        "submitted_at": utcnow_iso(),
        "question_type": q_type,
        "selected_option": int(data.selected_option) if data.selected_option is not None else None,
    }
    db.reference(f"/submissions/{sub_id}").set(submission)
    return {"message": "Submitted successfully", "submission_id": sub_id, "score": score} 
