"""
Student Routes
API endpoints for student operations — Firebase Realtime Database
"""

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from firebase_admin import db
from io import BytesIO
from app.routes.auth import get_current_user
from app.utils.rbac import has_permission

router = APIRouter(prefix="/api/student", tags=["student"])


def _language_label(value):
    labels = {
        "python": "Python",
        "py": "Python",
        "c": "C",
        "cpp": "C++",
        "c++": "C++",
        "java": "Java",
        "mcq": "MCQ",
    }
    key = str(value or "").strip().lower()
    return labels.get(key, value or "N/A")


def _date_only(value):
    if not value:
        return None
    return str(value)[:10]


def _safe_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _format_duration(seconds):
    if seconds is None:
        return "N/A"
    total = max(0, int(seconds))
    minutes = total // 60
    remaining = total % 60
    return f"{minutes}m {remaining}s"


def _download_headers(filename: str, content_type: str):
    return {
        "Content-Disposition": f'attachment; filename="{filename}"',
        "Content-Type": content_type,
    }


def _build_student_test_history(current_user: dict):
    student_id = current_user["id"]
    if not has_permission(current_user.get("role"), "attempt_test"):
        return []

    all_tests = db.reference("/tests").get() or {}
    all_questions = db.reference("/questions").get() or {}
    all_subs = db.reference("/submissions").get() or {}
    all_attempts = db.reference("/attempts").get() or {}

    questions_by_test = {}
    for qid, q in all_questions.items():
        test_id = q.get("test_id")
        if not test_id:
            continue
        questions_by_test.setdefault(test_id, {})[qid] = q

    grouped = {}
    for test_id, sessions in all_attempts.items():
        if not isinstance(sessions, dict):
            continue
        test = all_tests.get(test_id, {})
        for session_id, students in sessions.items():
            if not isinstance(students, dict):
                continue
            attempt = students.get(student_id)
            if not attempt:
                continue
            key = f"{test_id}:{session_id}"
            grouped.setdefault(key, {
                "test_id": test_id,
                "session_id": session_id,
                "test_name": test.get("title", "Unknown Test"),
                "created_at": attempt.get("submitted_at") or attempt.get("started_at"),
                "languages": [],
                "question_scores": {},
            })

    for sub in all_subs.values():
        if sub.get("student_id") != student_id:
            continue
        test_id = sub.get("test_id")
        if not test_id:
            continue
        test = all_tests.get(test_id, {})
        session_id = sub.get("session_id") or test.get("current_session_id") or "default"
        key = f"{test_id}:{session_id}"
        row = grouped.setdefault(key, {
            "test_id": test_id,
            "session_id": session_id,
            "test_name": test.get("title", "Unknown Test"),
            "created_at": sub.get("submitted_at"),
            "languages": [],
            "question_scores": {},
        })
        submitted_at = sub.get("submitted_at")
        if submitted_at and (not row.get("created_at") or submitted_at > row["created_at"]):
            row["created_at"] = submitted_at
        if sub.get("language"):
            row["languages"].append(sub.get("language"))
        if sub.get("question_id"):
            row["question_scores"][sub["question_id"]] = float(sub.get("score", 0) or 0)

    history = []
    for row in grouped.values():
        questions = questions_by_test.get(row["test_id"], {})
        total_marks = 0.0
        marks = 0.0
        for qid, q in questions.items():
            points = float(q.get("points", 0) or 0)
            total_marks += points
            score = row["question_scores"].get(qid, 0)
            marks += (score / 100.0) * points

        language_counts = {}
        for lang in row["languages"]:
            label = _language_label(lang)
            if label == "MCQ":
                continue
            language_counts[label] = language_counts.get(label, 0) + 1
        language = max(language_counts, key=language_counts.get) if language_counts else "N/A"

        history.append({
            "test_id": row["test_id"],
            "session_id": row["session_id"],
            "test_name": row["test_name"],
            "marks": round(marks, 2),
            "total_marks": round(total_marks, 2),
            "language": language,
            "date": _date_only(row.get("created_at")),
            "created_at": row.get("created_at"),
        })

    history.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    return history


@router.get("/test-history")
def get_test_history(current_user: dict = Depends(get_current_user)):
    """
    Returns one row per attempted test/session for the logged-in student.
    Firebase stores attempts under /attempts and submissions under /submissions;
    this endpoint presents the requested test_attempts-style shape.
    """
    return _build_student_test_history(current_user)


@router.get("/download-report/{test_id}")
def download_basic_report(test_id: str, session_id: str | None = None, current_user: dict = Depends(get_current_user)):
    """
    Basic legacy report download.
    Kept separate from /test-report/{test_id} so the quick download and
    detailed Excel report can fail/succeed independently.
    """
    report = _student_report_data(test_id, session_id, current_user)
    if not report:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="No report data found for this test attempt.")

    lines = [
        "SlashCoder Basic Report",
        f"Test Name,{report['test_name']}",
        f"Date,{report['date']}",
        f"Student Name,{report['student_name']}",
        f"Email / ID,{report['student_email']}",
        f"Language Used,{report['language']}",
        "",
        "Overall Summary",
    ]
    for key, value in report["summary"].items():
        lines.append(f"{key.replace('_', ' ').title()},{value}")

    lines.extend([
        "",
        "Q No,Question Name,Status,Marks,Language,Time",
    ])
    for q in report["questions"]:
        values = [q["q_no"], q["name"], q["status"], q["marks"], q["language"], q["time"]]
        lines.append(",".join(f'"{str(v).replace(chr(34), chr(34) + chr(34))}"' for v in values))

    lines.extend(["", "Test Case Summary", "Metric,Value"])
    for metric, value in report["test_cases"].items():
        lines.append(f"{metric},{value}")

    stream = BytesIO("\n".join(lines).encode("utf-8-sig"))
    return StreamingResponse(
        stream,
        media_type="text/csv",
        headers=_download_headers("SlashCoder_Report.csv", "text/csv"),
    )


def _student_report_data(test_id: str, session_id: str | None, current_user: dict):
    student_id = current_user["id"]
    if not has_permission(current_user.get("role"), "attempt_test"):
        return None

    test = db.reference(f"/tests/{test_id}").get()
    if not test:
        return None

    resolved_session_id = session_id or test.get("current_session_id") or "default"
    all_questions = db.reference("/questions").get() or {}
    all_test_cases = db.reference("/test_cases").get() or {}
    all_subs = db.reference("/submissions").get() or {}
    all_users = db.reference("/users").get() or {}
    attempt = db.reference(f"/attempts/{test_id}/{resolved_session_id}/{student_id}").get() or {}

    questions = [
        {"id": qid, **q}
        for qid, q in all_questions.items()
        if q.get("test_id") == test_id
    ]
    questions.sort(key=lambda q: q.get("created_at") or q.get("title") or "")

    student_subs = [
        {"id": sid, **sub}
        for sid, sub in all_subs.items()
        if sub.get("student_id") == student_id
        and sub.get("test_id") == test_id
        and (sub.get("session_id") or resolved_session_id) == resolved_session_id
    ]
    if not attempt and not student_subs:
        return None

    subs_by_question = {}
    for sub in student_subs:
        qid = sub.get("question_id")
        if qid:
            current = subs_by_question.get(qid)
            if not current or (sub.get("submitted_at") or "") > (current.get("submitted_at") or ""):
                subs_by_question[qid] = sub

    total_marks = sum(_safe_float(q.get("points")) for q in questions)
    earned_marks = 0.0
    correct = 0
    attempted = 0
    languages = {}
    total_test_cases = 0
    passed_test_cases = 0

    question_rows = []
    for index, q in enumerate(questions, start=1):
        sub = subs_by_question.get(q["id"])
        q_marks = _safe_float(q.get("points"))
        score = _safe_float(sub.get("score")) if sub else 0.0
        marks = round((score / 100.0) * q_marks, 2)
        earned_marks += marks
        if sub:
            attempted += 1
            lang_label = _language_label(sub.get("language"))
            if lang_label != "MCQ":
                languages[lang_label] = languages.get(lang_label, 0) + 1
            if score >= 100:
                correct += 1
            total = int(_safe_float(sub.get("total"), 0))
            passed = int(_safe_float(sub.get("passed"), 0))
        else:
            lang_label = "N/A"
            total = len([tc for tc in all_test_cases.values() if tc.get("question_id") == q["id"]])
            passed = 0
        total_test_cases += total
        passed_test_cases += passed
        status = "Correct" if sub and score >= 100 else "Attempted" if sub else "Not Attempted"
        question_rows.append({
            "q_no": index,
            "name": q.get("title") or f"Question {index}",
            "status": status,
            "marks": f"{marks:g} / {q_marks:g}",
            "language": lang_label,
            "time": f"{int(_safe_float(sub.get('execution_time_ms'), 0))}ms" if sub and sub.get("execution_time_ms") is not None else "N/A",
        })

    def aggregate_for_student(target_student_id: str):
        target_subs = [
            sub for sub in all_subs.values()
            if sub.get("student_id") == target_student_id
            and sub.get("test_id") == test_id
            and (sub.get("session_id") or resolved_session_id) == resolved_session_id
        ]
        by_question = {}
        for sub in target_subs:
            qid = sub.get("question_id")
            if qid and ((sub.get("submitted_at") or "") > (by_question.get(qid, {}).get("submitted_at") or "")):
                by_question[qid] = sub
        points = 0.0
        for q in questions:
            sub = by_question.get(q["id"])
            points += (_safe_float(sub.get("score")) / 100.0) * _safe_float(q.get("points")) if sub else 0.0
        return points

    participants = set()
    for sub in all_subs.values():
        if sub.get("test_id") == test_id and (sub.get("session_id") or resolved_session_id) == resolved_session_id and sub.get("student_id"):
            participants.add(sub.get("student_id"))
    participants.add(student_id)
    ranked = sorted(
        [{"student_id": sid, "marks": aggregate_for_student(sid)} for sid in participants],
        key=lambda x: x["marks"],
        reverse=True,
    )
    rank = next((idx + 1 for idx, row in enumerate(ranked) if row["student_id"] == student_id), 1)

    first_start = attempt.get("started_at")
    last_submit = attempt.get("submitted_at") or max((s.get("submitted_at") or "" for s in student_subs), default="")
    time_seconds = None
    if first_start and last_submit:
        from datetime import datetime
        try:
            start_dt = datetime.fromisoformat(first_start.replace("Z", "+00:00"))
            end_dt = datetime.fromisoformat(last_submit.replace("Z", "+00:00"))
            time_seconds = max(0, int((end_dt - start_dt).total_seconds()))
        except ValueError:
            time_seconds = None

    primary_language = max(languages, key=languages.get) if languages else "N/A"
    percentage = round((earned_marks / total_marks) * 100, 2) if total_marks > 0 else 0
    user = all_users.get(student_id, {})

    return {
        "test_name": test.get("title") or "Unknown Test",
        "date": _date_only(last_submit or first_start) or "N/A",
        "student_name": current_user.get("name") or user.get("name") or user.get("full_name") or "Unknown",
        "student_email": current_user.get("email") or user.get("email") or "Unknown",
        "language": primary_language,
        "summary": {
            "total_marks": f"{round(earned_marks, 2):g} / {round(total_marks, 2):g}",
            "percentage": f"{percentage}%",
            "rank": rank,
            "total_questions": len(questions),
            "attempted": attempted,
            "correct": correct,
            "time_taken": _format_duration(time_seconds),
        },
        "questions": question_rows,
        "test_cases": {
            "Total Test Cases": total_test_cases,
            "Passed": passed_test_cases,
            "Failed": max(0, total_test_cases - passed_test_cases),
        },
    }


def _build_report_workbook(report):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Report"

    dark_fill = PatternFill("solid", fgColor="1F4E78")
    section_fill = PatternFill("solid", fgColor="D9EAF7")
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    white_font = Font(color="FFFFFF", bold=True)
    bold_font = Font(bold=True)
    thin = Side(style="thin", color="B7C9D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    for col, width in {"A": 12, "B": 32, "C": 18, "D": 18, "E": 18, "F": 16}.items():
        ws.column_dimensions[col].width = width

    ws.merge_cells("A1:F1")
    ws["A1"] = "SlashCoder"
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].fill = dark_fill
    ws["A1"].alignment = center

    ws.merge_cells("A2:F2")
    ws["A2"] = "Test Performance Report"
    ws["A2"].font = Font(bold=True, size=14)
    ws["A2"].alignment = center

    ws["A4"] = "Test Name"
    ws["B4"] = report["test_name"]
    ws["D4"] = "Date"
    ws["E4"] = report["date"]

    ws.merge_cells("A6:F6")
    ws["A6"] = "Student Details"
    ws["A6"].font = bold_font
    ws["A6"].fill = section_fill
    ws["A7"], ws["B7"] = "Name", report["student_name"]
    ws["A8"], ws["B8"] = "Email / ID", report["student_email"]
    ws["A9"], ws["B9"] = "Language Used", report["language"]

    ws.merge_cells("A11:F11")
    ws["A11"] = "Overall Summary"
    ws["A11"].font = bold_font
    ws["A11"].fill = section_fill
    summary_items = list(report["summary"].items())
    for idx, (label, value) in enumerate(summary_items, start=12):
        ws[f"A{idx}"] = label.replace("_", " ").title()
        ws[f"B{idx}"] = value

    table_start = 21
    ws.merge_cells(start_row=table_start, start_column=1, end_row=table_start, end_column=6)
    ws.cell(table_start, 1, "Question-wise Table")
    ws.cell(table_start, 1).font = bold_font
    ws.cell(table_start, 1).fill = section_fill

    headers = ["Q No", "Question Name", "Status", "Marks", "Language", "Time"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(table_start + 1, col, header)
        cell.font = bold_font
        cell.fill = header_fill
        cell.alignment = center

    for row_idx, q in enumerate(report["questions"], start=table_start + 2):
        values = [q["q_no"], q["name"], q["status"], q["marks"], q["language"], q["time"]]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row_idx, col_idx, value)

    tc_start = table_start + 4 + len(report["questions"])
    ws.merge_cells(start_row=tc_start, start_column=1, end_row=tc_start, end_column=2)
    ws.cell(tc_start, 1, "Test Case Summary")
    ws.cell(tc_start, 1).font = bold_font
    ws.cell(tc_start, 1).fill = section_fill
    ws.cell(tc_start + 1, 1, "Metric")
    ws.cell(tc_start + 1, 2, "Value")
    ws.cell(tc_start + 1, 1).font = bold_font
    ws.cell(tc_start + 1, 2).font = bold_font
    ws.cell(tc_start + 1, 1).fill = header_fill
    ws.cell(tc_start + 1, 2).fill = header_fill
    for idx, (metric, value) in enumerate(report["test_cases"].items(), start=tc_start + 2):
        ws.cell(idx, 1, metric)
        ws.cell(idx, 2, value)

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = border
            cell.alignment = left if cell.column in {2, 3} else center

    ws.freeze_panes = "A22"
    return wb


@router.get("/test-report/{test_id}")
def download_test_report(test_id: str, session_id: str | None = None, current_user: dict = Depends(get_current_user)):
    report = _student_report_data(test_id, session_id, current_user)
    if not report:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="No report data found for this test attempt.")

    workbook = _build_report_workbook(report)
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    filename = "SlashCoder_Report.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/my-submissions")
def get_my_submissions(current_user: dict = Depends(get_current_user)):
    """
    Returns all submissions made by the currently logged-in student.
    Joins with /tests to include the test title.
    """
    student_id = current_user["id"]
    if not has_permission(current_user.get("role"), "attempt_test"):
        return []

    all_subs = db.reference("/submissions").get() or {}
    all_tests = db.reference("/tests").get() or {}

    result = []

    for sub_id, sub in all_subs.items():
        if sub.get("student_id") != student_id:
            continue

        test_id = sub.get("test_id")

        # Only show submissions from the current session.
        test = all_tests.get(test_id, {})
        current_session_id = test.get("current_session_id")
        if current_session_id and sub.get("session_id") != current_session_id:
            continue

        result.append({
            "submission_id": sub_id,
            "test_id": test_id,
            "test_title": test.get("title", "Unknown Test"),
            "score": sub.get("score", 0),
            "language": sub.get("language"),
            "submitted_at": sub.get("submitted_at"),
            "session_id": sub.get("session_id"),
        })

    # Sort by submitted_at descending.
    result.sort(key=lambda x: x.get("submitted_at") or "", reverse=True)

    # Deduplicate by test_id — keep only the latest submission per test.
    seen = set()
    deduped = []
    for item in result:
        if item["test_id"] not in seen:
            seen.add(item["test_id"])
            deduped.append(item)

    return deduped
